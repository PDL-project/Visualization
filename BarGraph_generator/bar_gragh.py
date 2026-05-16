#!/usr/bin/env python3
"""
plot_map_thor.py
Generates a grouped bar chart for the MAP-THOR environment.
Visual style matches base.png.
"""

import re
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from pathlib import Path
import openpyxl

EXCEL_PATH = Path(__file__).parent / 'Robot_Algorithm_Performance.xlsx'
OUTPUT_PATH = Path(__file__).parent / 'MAP-THOR_comparison.png'
ENV_TARGET = 'MAP-THOR'


# ── helpers ───────────────────────────────────────────────────────────────────

def _to_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(',', '')
    s = re.sub(r's$', '', s)
    try:
        return float(s)
    except ValueError:
        return None


def _direction(text):
    t = str(text)
    if '↑' in t or '(↑)' in t or '(up)' in t.lower():
        return '↑'
    if '↓' in t or '(↓)' in t or '(down)' in t.lower():
        return '↓'
    return ''


def _strip_markers(text):
    return re.sub(r'\s*[\(\（][↑↓↑↓▲▼][\)\）]\s*', '', str(text)).strip()


def _fmt_val(val):
    if val is None:
        return ''
    if val >= 1e6:  return f'{val/1e6:.2f}M'
    if val >= 1e4:  return f'{val/1e3:.1f}K'
    if val >= 100:  return f'{val:.0f}'
    if val >= 10:   return f'{val:.1f}'
    if val >= 1:    return f'{val:.2f}'
    return f'{val:.3f}'


def _abbrev(name):
    clean = re.sub(r'\s*\([^)]*\)', '', name).strip()
    if len(clean) <= 6:
        return clean
    return ''.join(w[0].upper() for w in clean.split())


# ── parser ────────────────────────────────────────────────────────────────────

def parse_map_thor(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # find MAP-THOR column range from merged cells in row 1
    env_col = None
    for m in ws.merged_cells.ranges:
        if m.min_row == 1:
            val = ws.cell(1, m.min_col).value
            if val == ENV_TARGET:
                env_col = (m.min_col, m.max_col)
                break

    if env_col is None:
        raise ValueError(f'"{ENV_TARGET}" not found in row 1.')

    col_min, col_max = env_col

    # metrics from row 2
    metrics = []
    for c in range(col_min, col_max + 1):
        raw = ws.cell(2, c).value
        if raw is not None:
            metrics.append((_strip_markers(raw), _direction(raw), str(raw)))

    # algorithms + values from row 3+
    algorithms = []
    data = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        algo = row[0]
        if algo is None:
            continue
        algorithms.append(str(algo))
        vals = [_to_float(row[c - 1]) for c in range(col_min, col_min + len(metrics))]
        data[str(algo)] = vals

    return metrics, algorithms, data


# ── plotter ───────────────────────────────────────────────────────────────────

def plot():
    metrics, algorithms, data = parse_map_thor(EXCEL_PATH)
    n_met = len(metrics)
    n_alg = len(algorithms)

    proposed_idx = next(
        (i for i, a in enumerate(algorithms) if 'proposed' in a.lower()), 0
    )

    print(f'Metrics ({n_met}): {[m[0] for m in metrics]}')
    print(f'Algorithms ({n_alg}): {algorithms}')
    print(f'Proposed: {algorithms[proposed_idx]}')

    # ── rcParams (publication quality) ────────────────────────────────────────
    plt.rcParams.update({
        'font.family':     'sans-serif',
        'font.size':        9,
        'axes.labelsize':   9,
        'xtick.labelsize':  8,
        'ytick.labelsize':  8,
        'legend.fontsize':  9,
        'axes.linewidth':   0.8,
        'grid.linewidth':   0.4,
    })

    # ── colors ────────────────────────────────────────────────────────────────
    palette = ['#2E75B6', '#548235', '#C00000', '#ED7D31',
               '#7030A0', '#7F7F7F', '#00B0F0', '#FF0000']
    algo_color = {a: palette[i % len(palette)] for i, a in enumerate(algorithms)}

    # ── bar / group geometry ──────────────────────────────────────────────────
    bar_w     = 0.09
    group_w   = n_alg * bar_w
    group_gap = 0.09
    step      = group_w + group_gap
    group_cx  = np.arange(n_met) * step
    bar_off   = np.arange(n_alg) * bar_w - (n_alg - 1) * bar_w / 2

    x_min = group_cx[0]  - group_w / 2 - group_gap / 2
    x_max = group_cx[-1] + group_w / 2 + group_gap / 2

    # ── figure (single axes) ──────────────────────────────────────────────────
    # figure width scaled so bars keep same physical size as step shrinks
    fig_w = 6.4 * (step / 0.36)
    fig, ax = plt.subplots(figsize=(fig_w, 5.5), facecolor='white')
    fig.subplots_adjust(left=0.11, right=0.97, top=0.90, bottom=0.13)

    # ── draw bars ─────────────────────────────────────────────────────────────
    for m_idx in range(n_met):
        for a_idx, algo in enumerate(algorithms):
            val = data[algo][m_idx]
            if val is None:
                continue
            x = group_cx[m_idx] + bar_off[a_idx]
            ax.bar(x, val, width=bar_w,
                   color=algo_color[algo],
                   edgecolor='#333333', linewidth=0.4,
                   label=algo if m_idx == 0 else '_nolegend_')

    # ── axes style ────────────────────────────────────────────────────────────
    ax.set_xlim(x_min, x_max)
    ax.set_yscale('log')
    ax.set_facecolor('white')
    ax.tick_params(which='major', left=True, labelleft=True,
                   labelsize=7, length=5, width=0.6, direction='in')
    ax.tick_params(which='minor', left=False, labelleft=False,
                   length=0, width=0)
    ax.tick_params(which='both', bottom=False, top=False, right=False)
    ax.yaxis.set_major_locator(matplotlib.ticker.LogLocator(base=10.0, numticks=6))
    ax.yaxis.set_major_formatter(matplotlib.ticker.LogFormatterMathtext())
    ax.minorticks_off()
    ax.yaxis.grid(False)
    ax.set_axisbelow(True)
    for spine in ax.spines.values():
        spine.set_color('#333333')
        spine.set_linewidth(0.6)

    # ── x-axis metric labels ──────────────────────────────────────────────────
    ax.set_xticks(group_cx)
    ax.set_xticklabels(
        [f'{_abbrev(m[0])} ({m[1]})' if m[1] else _abbrev(m[0]) for m in metrics],
        fontsize=8.5
    )
    ax.tick_params(axis='x', which='major', bottom=False, labelbottom=True)

    # ── global y-limits ───────────────────────────────────────────────────────
    all_vals = [data[a][m] for a in algorithms for m in range(n_met)
                if data[a][m] is not None and data[a][m] > 0]
    if all_vals:
        ax.set_ylim(min(all_vals) * 0.88, max(all_vals) * 3.5)

    # ── ratio arrows ──────────────────────────────────────────────────────────
    if n_alg >= 2:
        for m_idx in range(n_met):
            m_dir = metrics[m_idx][1]
            prop_val = data[algorithms[proposed_idx]][m_idx]
            if prop_val is None or prop_val <= 0:
                continue
            worst_val, worst_a_idx = None, None
            for a_idx, algo in enumerate(algorithms):
                if a_idx == proposed_idx:
                    continue
                val = data[algo][m_idx]
                if val is None or val <= 0:
                    continue
                if m_dir == '↑':
                    if worst_val is None or val < worst_val:
                        worst_val, worst_a_idx = val, a_idx
                else:
                    if worst_val is None or val > worst_val:
                        worst_val, worst_a_idx = val, a_idx
            if worst_val is None or worst_val == prop_val:
                continue
            lo_val = min(prop_val, worst_val)
            hi_val = max(prop_val, worst_val)
            ratio  = hi_val / lo_val
            if ratio < 2.0:
                continue
            lo_a_idx = proposed_idx if prop_val < worst_val else worst_a_idx
            arr_x = group_cx[m_idx] + bar_off[lo_a_idx]
            ax.annotate('',
                        xy=(arr_x, hi_val), xytext=(arr_x, lo_val),
                        arrowprops=dict(arrowstyle='<->', color='#C00000',
                                        lw=0.5, mutation_scale=7))
            label = f'{ratio:.1f}×' if ratio < 10 else f'{ratio:.0f}×'
            ax.text(arr_x, hi_val * 1.15, label,
                    ha='center', va='bottom',
                    fontsize=7.5, color='#C00000', fontweight='bold')

    # ── environment label ─────────────────────────────────────────────────────
    fig.text(0.54, 0.04, ENV_TARGET,
             ha='center', va='bottom', fontsize=10, fontweight='bold',
             color='#111111', transform=fig.transFigure)

    # ── legend ────────────────────────────────────────────────────────────────
    handles = [mpatches.Patch(color=algo_color[a], label=a) for a in algorithms]
    fig.legend(handles=handles, loc='upper center', ncol=n_alg,
               fontsize=9, frameon=False, bbox_to_anchor=(0.5, 1.0),
               handlelength=1.2, handletextpad=0.5, columnspacing=1.0)

    # ── save ──────────────────────────────────────────────────────────────────
    plt.savefig(OUTPUT_PATH, dpi=300, bbox_inches='tight', facecolor='white')
    print(f'Saved → {OUTPUT_PATH}')
    plt.show()


if __name__ == '__main__':
    plot()
